"""Excel 批量导入服务 (组织/员工/设备).

- 模板下载 / XLSX 导入 / 批次记录 / 行级校验 / 部分成功 / 失败行下载 / 幂等
- 采用流式逐行处理 (openpyxl read_only), 不在内存一次性装载超大文件
"""

from __future__ import annotations

import io
import uuid
from datetime import date

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import AppError
from app.models.device import Device
from app.models.imports import ImportBatch, ImportItem
from app.models.org import Employee, OrganizationNode, Store
from app.services.security_context import TenantContext

TEMPLATES: dict[str, tuple[list[str], list[str]]] = {
    "ORGANIZATION": (
        ["node_type", "name", "code", "parent_code", "sort_order"],
        ["node_type(总部HQ/区域REGION/门店STORE)", "节点名称", "节点编码", "上级编码(可空)", "排序"],
    ),
    "EMPLOYEE": (
        ["employee_no", "name", "mobile", "job_title", "store_code", "joined_at"],
        ["员工号", "姓名", "手机号", "岗位", "门店编码", "入职日期(YYYY-MM-DD)"],
    ),
    "DEVICE": (
        ["device_code", "device_type", "vendor", "model"],
        ["设备码(SN)", "设备类型(BADGE)", "厂商", "型号"],
    ),
}


class ImportService:
    def __init__(self, session: AsyncSession, ctx: TenantContext) -> None:
        self.session = session
        self.ctx = ctx

    @staticmethod
    def build_template(import_type: str) -> bytes:
        keys, headers = TEMPLATES[import_type]
        wb = Workbook()
        ws = wb.active
        ws.title = "导入模板"
        ws.append(headers)
        ws.append([""] * len(keys))
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def run_import(
        self, import_type: str, file_name: str, content: bytes
    ) -> ImportBatch:
        if import_type not in TEMPLATES:
            raise AppError(400, "invalid_import_type", "不支持的导入类型")
        keys, _ = TEMPLATES[import_type]
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AppError(400, "invalid_xlsx", "无法解析 XLSX 文件") from exc
        ws = wb.active

        batch = ImportBatch(
            tenant_id=self.ctx.tenant_id,
            import_type=import_type,
            file_name=file_name,
            status="PROCESSING",
            created_by=self.ctx.user.id if self.ctx.user else None,
        )
        self.session.add(batch)
        await self.session.flush()

        total = success = failed = 0
        errors: list[str] = []
        row_no = 0
        for row in ws.iter_rows(values_only=True):
            row_no += 1
            if row_no == 1:
                continue  # 表头
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            total += 1
            raw = {keys[i]: ("" if i >= len(row) or row[i] is None else str(row[i]).strip()) for i in range(len(keys))}
            try:
                await self._apply_row(import_type, raw)
                success += 1
                self.session.add(
                    ImportItem(
                        tenant_id=self.ctx.tenant_id,
                        batch_id=batch.id,
                        row_no=row_no,
                        status="SUCCESS",
                        raw_data=raw,
                        legacy_key=raw.get("employee_no") or raw.get("device_code") or raw.get("code"),
                    )
                )
            except Exception as exc:  # noqa: BLE001
                failed += 1
                msg = str(exc)
                errors.append(msg[:200])
                self.session.add(
                    ImportItem(
                        tenant_id=self.ctx.tenant_id,
                        batch_id=batch.id,
                        row_no=row_no,
                        status="FAILED",
                        raw_data=raw,
                        error_message=msg[:1000],
                    )
                )
        wb.close()

        batch.total_rows = total
        batch.success_rows = success
        batch.failed_rows = failed
        batch.status = "SUCCEEDED" if failed == 0 else ("PARTIAL" if success > 0 else "FAILED")
        if errors:
            batch.error_summary = "; ".join(errors[:5])
        await self.session.commit()
        await self.session.refresh(batch)
        return batch

    # ---- 行处理 (幂等: 已存在同键记录则跳过) ----
    async def _apply_row(self, import_type: str, raw: dict[str, str]) -> None:
        tid = self.ctx.tenant_id
        if import_type == "EMPLOYEE":
            emp_no = raw.get("employee_no") or ""
            if not emp_no:
                raise ValueError("员工号不能为空")
            existing = await self.session.scalar(
                select(Employee).where(
                    Employee.tenant_id == tid, Employee.employee_no == emp_no
                )
            )
            if existing is not None:
                return  # 幂等: 已存在
            mobile = raw.get("mobile") or ""
            if not mobile:
                raise ValueError("手机号不能为空")
            if len(mobile) < 5:
                raise ValueError("手机号格式不正确")
            store_id = None
            if raw.get("store_code"):
                store = await self.session.scalar(
                    select(Store).where(Store.tenant_id == tid, Store.code == raw["store_code"])
                )
                if store is None:
                    raise ValueError(f"门店编码不存在: {raw['store_code']}")
                store_id = store.id
            joined = None
            if raw.get("joined_at"):
                try:
                    joined = date.fromisoformat(raw["joined_at"])
                except ValueError as exc:
                    raise ValueError(f"入职日期格式错误: {raw['joined_at']}") from exc
            self.session.add(
                Employee(
                    tenant_id=tid,
                    employee_no=emp_no,
                    name=raw.get("name") or "",
                    mobile=mobile,
                    job_title=raw.get("job_title") or None,
                    store_id=store_id,
                    joined_at=joined,
                )
            )
        elif import_type == "DEVICE":
            code = raw.get("device_code") or ""
            if not code:
                raise ValueError("设备码不能为空")
            existing = await self.session.scalar(
                select(Device).where(Device.tenant_id == tid, Device.device_code == code)
            )
            if existing is not None:
                return  # 幂等
            self.session.add(
                Device(
                    tenant_id=tid,
                    device_code=code,
                    device_type=raw.get("device_type") or "BADGE",
                    vendor=raw.get("vendor") or None,
                    model=raw.get("model") or None,
                )
            )
        elif import_type == "ORGANIZATION":
            code = raw.get("code") or ""
            if not code:
                raise ValueError("节点编码不能为空")
            existing = await self.session.scalar(
                select(OrganizationNode).where(OrganizationNode.tenant_id == tid, OrganizationNode.code == code)
            )
            if existing is not None:
                return  # 幂等
            node_type = raw.get("node_type") or "GROUP"
            if node_type not in ("HQ", "REGION", "STORE", "GROUP"):
                raise ValueError(f"非法节点类型: {node_type}")
            parent_id = None
            if raw.get("parent_code"):
                parent = await self.session.scalar(
                    select(OrganizationNode).where(
                        OrganizationNode.tenant_id == tid, OrganizationNode.code == raw["parent_code"]
                    )
                )
                if parent is None:
                    raise ValueError(f"上级编码不存在: {raw['parent_code']}")
                parent_id = parent.id
            self.session.add(
                OrganizationNode(
                    tenant_id=tid,
                    node_type=node_type,
                    name=raw.get("name") or code,
                    code=code,
                    parent_id=parent_id,
                    sort_order=int(raw.get("sort_order") or 0),
                )
            )
        else:  # pragma: no cover
            raise AppError(400, "invalid_import_type", "不支持的导入类型")

    async def failure_workbook(self, batch_id: uuid.UUID) -> bytes:
        batch = await self.session.get(ImportBatch, batch_id)
        if batch is None or str(batch.tenant_id) != str(self.ctx.tenant_id):
            raise AppError(404, "not_found", "导入批次不存在")
        items = (
            (
                await self.session.execute(
                    select(ImportItem).where(
                        ImportItem.batch_id == batch_id, ImportItem.status == "FAILED"
                    )
                )
            )
            .scalars()
            .all()
        )
        keys, headers = TEMPLATES[batch.import_type]
        wb = Workbook()
        ws = wb.active
        ws.title = "失败明细"
        ws.append([*headers, "失败原因"])
        for item in items:
            raw = item.raw_data or {}
            ws.append([raw.get(k, "") for k in keys] + [item.error_message or ""])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
