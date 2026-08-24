"""Excel 导入测试: 部分成功 / 幂等 / 失败行下载."""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from app.modules.imports.service import TEMPLATES, ImportService
from tests.conftest import build_org


def _xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestEmployeeImport:
    async def test_partial_success(self, db, session_factory) -> None:
        org = await build_org(session_factory)
        # 需要 ctx 对象; 通过 API 更贴近真实, 但此处直接构造 service
        from app.services.security_context import TenantContext

        user = None
        ctx = TenantContext(user=user, tenant_id=org["tenant"].id, is_super_admin=True)  # type: ignore[arg-type]
        service = ImportService(db, ctx)
        keys, _ = TEMPLATES["EMPLOYEE"]
        rows = [
            keys,  # 表头
            ["E001", "张三", "13900000001", "营业员", "S-A", "2026-01-01"],
            ["E002", "李四", "13900000002", "营业员", "S-A", "2026-02-01"],
            ["E003", "王五", "123", "营业员", "S-A", "2026-03-01"],  # 手机号非法(过短)
            ["E004", "赵六", "13900000004", "营业员", "S-NOT-EXIST", "2026-04-01"],  # 门店不存在
        ]
        batch = await service.run_import("EMPLOYEE", "employees.xlsx", _xlsx(rows))
        assert batch.status == "PARTIAL"
        assert batch.total_rows == 4
        assert batch.success_rows == 2
        assert batch.failed_rows == 2
        # 成功员工已落库
        from sqlalchemy import select

        from app.models.org import Employee

        emp = await db.scalar(select(Employee).where(Employee.employee_no == "E001"))
        assert emp is not None
        assert emp.store_id is not None

    async def test_idempotent_reimport(self, db, session_factory) -> None:
        org = await build_org(session_factory)
        from app.services.security_context import TenantContext

        ctx = TenantContext(user=None, tenant_id=org["tenant"].id, is_super_admin=True)  # type: ignore[arg-type]
        service = ImportService(db, ctx)
        keys, _ = TEMPLATES["EMPLOYEE"]
        rows = [keys, ["E100", "重复", "13900000100", "营业员", "", ""]]
        b1 = await service.run_import("EMPLOYEE", "a.xlsx", _xlsx(rows))
        b2 = await service.run_import("EMPLOYEE", "a.xlsx", _xlsx(rows))
        assert b1.success_rows == 1
        assert b2.success_rows == 1  # 幂等: 已存在员工跳过, 不重复
        from sqlalchemy import func, select

        from app.models.org import Employee

        total = await db.scalar(
            select(func.count()).select_from(Employee).where(Employee.employee_no == "E100")
        )
        assert total == 1

    async def test_failure_download(self, db, session_factory) -> None:
        org = await build_org(session_factory)
        from app.services.security_context import TenantContext

        ctx = TenantContext(user=None, tenant_id=org["tenant"].id, is_super_admin=True)  # type: ignore[arg-type]
        service = ImportService(db, ctx)
        keys, _ = TEMPLATES["EMPLOYEE"]
        batch = await service.run_import(
            "EMPLOYEE", "bad.xlsx", _xlsx([keys, ["X1", "坏", "nope", "", "", ""]])
        )
        content = await service.failure_workbook(batch.id)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.max_row == 2  # 表头 + 1 失败行
        assert "失败原因" in [c.value for c in ws[1]]
